#!/usr/bin/env python3
"""為 Shopline 產品匯出加入智能分類（專業／詳細／功能／針對／類型）。"""

from __future__ import annotations

import argparse
import csv
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import unquote, urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
_UA = "SiteSpider-ShoplineClassify/1.0"
_SKIP_CATS = {"全部商品", "散貨佬精選優惠", "近期至抵", "瘦身精選"}

# slug 或顯示名 → (專業分類, 詳細分類)
_TAXONOMY: dict[str, tuple[str, str]] = {
    # 保健
    "個人保健專區": ("營養保健", "綜合保健"),
    "護肝": ("營養保健", "護肝排毒"),
    "益生菌": ("營養保健", "益生菌／腸道"),
    "減肥產品": ("營養保健", "纖體瘦身"),
    "維他命": ("營養保健", "維他命／礦物質"),
    "護眼": ("營養保健", "護眼明目"),
    "男士專區": ("營養保健", "男士保健"),
    "女士專區": ("營養保健", "女士保健"),
    "兒童專區": ("營養保健", "兒童保健"),
    "兒童成長發展": ("營養保健", "兒童成長"),
    "心血管健康": ("營養保健", "心血管"),
    "腦部發展": ("營養保健", "腦部／記憶"),
    "藥油鐵打": ("營養保健", "外用的藥油／跌打"),
    "補鈣": ("營養保健", "鈣質／骨骼"),
    "祛濕排毒": ("營養保健", "祛濕排毒"),
    "美肌飲": ("營養保健", "美容飲品／膠原蛋白"),
    "日本藥品": ("日本藥品", "日本 OTC／保健"),
    "品牌專區（保健）": ("品牌專區", "保健品牌"),
    "blackmores": ("品牌專區", "Blackmores"),
    "swisse": ("品牌專區", "Swisse"),
    "proven": ("品牌專區", "Proven"),
    "return-回本": ("品牌專區", "Return 回本"),
    "bg-pro": ("品牌專區", "BG Pro"),
    "維特健靈": ("品牌專區", "維特健靈"),
    "御藥堂": ("品牌專區", "御藥堂"),
    "森下仁丹": ("品牌專區", "森下仁丹"),
    "樂道": ("品牌專區", "樂道 Noto"),
    "愛完全": ("品牌專區", "愛完全 WholeLove"),
    "紐西蘭正品": ("品牌專區", "紐西蘭正品"),
    # 護膚
    "護膚品專區": ("護膚美容", "綜合護膚"),
    "洗面奶": ("護膚美容", "潔面／洗面"),
    "爽膚水": ("護膚美容", "爽膚／化妝水"),
    "精華": ("護膚美容", "精華液"),
    "面膜": ("護膚美容", "面膜"),
    "眼部護理": ("護膚美容", "眼部護理"),
    "面霜": ("護膚美容", "面霜／乳霜"),
    "美白產品": ("護膚美容", "美白淡斑"),
    "精選護膚套裝": ("護膚美容", "護膚套裝"),
    "痘痘護理": ("護膚美容", "痘痘／暗瘡"),
    "唇部護理": ("護膚美容", "唇部護理"),
    # 彩妝
    "彩妝專區": ("彩妝造型", "綜合彩妝"),
    "妝前隔離": ("彩妝造型", "妝前／隔離"),
    "粉底": ("彩妝造型", "粉底"),
    "遮瑕膏": ("彩妝造型", "遮瑕"),
    "定妝粉": ("彩妝造型", "定妝"),
    "眼影": ("彩妝造型", "眼影"),
    "唇膏": ("彩妝造型", "唇膏／唇彩"),
    "化妝工具": ("彩妝造型", "化妝工具"),
    # 個人護理
    "個人護理": ("個人護理", "綜合個護"),
    "卸妝": ("個人護理", "卸妝"),
    "護手霜": ("個人護理", "手部護理"),
    "身體乳": ("個人護理", "身體護理"),
    "沐浴露": ("個人護理", "沐浴"),
    "私密護理": ("個人護理", "私密護理"),
    "防曬": ("個人護理", "防曬"),
    "頭髮護理": ("個人護理", "頭髮護理"),
    "口腔護理": ("個人護理", "口腔護理"),
    # 香水
    "香水專區": ("香水香氛", "香水"),
    "gucci-1": ("香水香氛", "Gucci"),
    "lanvin": ("香水香氛", "Lanvin"),
    "香水試用裝": ("香水香氛", "試香／小樣"),
    # 家居／生活
    "家居用品": ("家居生活", "家居清潔"),
    "洗衫用品": ("家居生活", "洗衣用品"),
    "情趣用品": ("家居生活", "成人用品"),
    "精選生活小物": ("家居生活", "生活小物"),
    "玩具專區": ("休閒娛樂", "玩具／娛樂"),
    "美食專區": ("食品美食", "零食／食品"),
}

_PARENT_ONLY = {
    "個人保健專區",
    "護膚品專區",
    "彩妝專區",
    "個人護理",
    "香水專區",
    "家居用品",
    "品牌專區（保健）",
}

# (關鍵字, 標籤) — 依優先序
_FUNCTION_RULES: list[tuple[str, str]] = [
    (r"益生菌|益菌|lgg|bifido|lactobac", "益生菌／腸道"),
    (r"護肝|淨肝|乳薊|水飛薊|肝", "護肝"),
    (r"補鈣|鈣片|珊瑚鈣|caltrate|骨骼", "補鈣／骨骼"),
    (r"膠原|collagen|美肌飲", "膠原蛋白／美肌"),
    (r"免疫|igG|免疫球", "免疫調節"),
    (r"維他命|綜合維他|multivit|ultivite", "綜合維他命"),
    (r"魚油|omega|DHA|EPA", "魚油／Omega-3"),
    (r"葉黃素|護眼|blue.?light|山桑子", "護眼"),
    (r"腦|記憶|銀杏|DHA|腦活素|專注", "腦部／記憶"),
    (r"鐵|補血|貧血|holdbody.?iron", "補鐵／補血"),
    (r"纖體|瘦身|減肥|代餐|排脂|svelty", "纖體瘦身"),
    (r"祛濕|除濕|竹酢|吸濕|排毒", "祛濕排毒"),
    (r"關節|腰腎|骨痛|透骨|黑玉貼|鎮痛", "關節／筋骨"),
    (r"Q10|輔酶|coenzyme", "輔酶 Q10"),
    (r"美白|亮白|淡斑|whitening|煙酰胺|熊果", "美白"),
    (r"保濕|補水|玻尿酸|透明質酸|hyaluron|b5", "保濕補水"),
    (r"抗敏|舒敏|敏感|avène|活泉水", "舒緩抗敏"),
    (r"緊緻|抗老|皺紋|逆齡|collagen", "抗老緊緻"),
    (r"去痘|暗瘡|痘痘|acne", "控油去痘"),
    (r"防曬|spf|sun", "防曬"),
    (r"育髮|生髮|掉髮|頭皮|洗髮", "頭皮／育髮"),
    (r"卸妝|cleans", "卸妝清潔"),
    (r"口腔|牙膏|漱口水|牙", "口腔護理"),
    (r"助眠|睡眠|甜睡", "助眠"),
    (r"心血管|血壓|胆固醇", "心血管"),
    (r"消石|腎|泌尿", "泌尿／消石"),
    (r"口罩|防疫", "防護口罩"),
    (r"驅蟲|防蠅|排水", "家居防蟲"),
]

_TARGET_RULES: list[tuple[str, str]] = [
    (r"兒童|kid|child|小童|嬰|baby", "兒童"),
    (r"孕婦|懷孕|maternity|媽媽鈣", "孕婦／哺乳期"),
    (r"男士|男性|male|men'?s", "男士"),
    (r"女士|女性|female|women", "女士"),
    (r"熟齡|銀髮|長者|40代|50代", "熟齡"),
    (r"敏感|抗敏|低敏", "敏感肌"),
    (r"油性|控油|痘痘", "油性／問題肌"),
    (r"乾性|缺水|脫皮", "乾燥肌"),
    (r"全效|全家|family", "全家適用"),
]

_TYPE_RULES: list[tuple[str, str]] = [
    (r"膠囊|capsule|粒裝|粒\b|片裝|片\b|丸", "膠囊／片劑"),
    (r"粉|蛋白粉|沖劑|顆粒", "粉末／沖劑"),
    (r"飲|ml x|毫升|支裝|美肌飲|口服液", "飲品／口服液"),
    (r"面膜|mask|片裝", "面膜"),
    (r"爽膚|化妝水|toner", "爽膚水"),
    (r"精華|serum|essence", "精華"),
    (r"乳液|乳霜|面霜|cream|lotion", "乳液／面霜"),
    (r"眼霜", "眼霜"),
    (r"洗面|潔面|cleanser", "潔面乳"),
    (r"防曬|sunscreen|spf", "防曬乳"),
    (r"洗髮|shampoo", "洗髮"),
    (r"護髮|conditioner|髮膜", "護髮"),
    (r"沐浴|body wash|沐浴露", "沐浴露"),
    (r"身體乳|body lotion", "身體乳"),
    (r"護手|hand cream", "護手霜"),
    (r"牙膏|漱口水|牙", "口腔用品"),
    (r"粉底|foundation", "粉底"),
    (r"遮瑕|concealer", "遮瑕"),
    (r"唇膏|lip", "唇膏"),
    (r"眼影|eye shadow", "眼影"),
    (r"香水|parfum|eau de", "香水"),
    (r"貼\b|patch|膏\b|油\b|藥油", "外用膏／貼／油"),
    (r"工具|梳|brush|氣墊梳", "工具／配件"),
    (r"口罩", "口罩"),
    (r"零食|食品|糖|餅|麵", "食品"),
]

# 關鍵字 fallback → (專業, 詳細)
_KEYWORD_TAXONOMY: list[tuple[str, str, str]] = [
    (r"益生菌|culturelle|g-niib|life.?space", "營養保健", "益生菌／腸道"),
    (r"護肝|淨肝|乳薊|御藥堂", "營養保健", "護肝排毒"),
    (r"鈣片|補鈣|caltrate|gnc.*鈣", "營養保健", "鈣質／骨骼"),
    (r"維他命|swisse|blackmores|gnc|bg.?pro|樂道|noto|維特健靈|森下仁丹|return|愛完全", "營養保健", "綜合保健"),
    (r"面膜|abib|ahc.*面膜", "護膚美容", "面膜"),
    (r"洗面|cleanser|潔面", "護膚美容", "潔面／洗面"),
    (r"爽膚|化妝水|toner", "護膚美容", "爽膚／化妝水"),
    (r"精華|serum", "護膚美容", "精華液"),
    (r"眼霜|眼膜|眼部", "護膚美容", "眼部護理"),
    (r"面霜|乳液|cream|lotion", "護膚美容", "面霜／乳霜"),
    (r"粉底|遮瑕|眼影|唇膏|彩妝", "彩妝造型", "綜合彩妝"),
    (r"洗髮|護髮|育髮|50惠|髮", "個人護理", "頭髮護理"),
    (r"沐浴|body wash", "個人護理", "沐浴"),
    (r"防曬|spf", "個人護理", "防曬"),
    (r"牙膏|口腔|漱口水", "個人護理", "口腔護理"),
    (r"香水|gucci|lanvin", "香水香氛", "香水"),
    (r"口罩", "家居生活", "防護口罩"),
    (r"洗衫|洗衣|排水|家居", "家居生活", "家居清潔"),
    (r"零食|食品|糖|餅|麵|奶粉|羊奶|美素", "食品美食", "奶粉／食品"),
    (r"梳|brush|氣墊|aveda", "個人護理", "美髮工具"),
    (r"牛初乳|蟲草|cs-4|catalo", "營養保健", "免疫／滋補"),
    (r"葉酸|蛋白粉|holdbody|免疫球|united med", "營養保健", "綜合保健"),
    (r"nmn|逆齡", "營養保健", "抗老／NMN"),
    (r"魚油|omega|globalab|optic", "營養保健", "魚油／護眼"),
    (r"關節|鎮痛|乳膏|seisyo|半月板", "營養保健", "關節／筋骨"),
    (r"胃藥|nexium|念慈|京都|nin jiom", "日本藥品", "日本 OTC／保健"),
    (r"避孕|evra", "個人護理", "私密護理"),
    (r"柔膚水|kiehl|fresh|保濕乳", "護膚美容", "爽膚／乳霜"),
    (r"溶脂|塑形|ppaebar|瘦腿|美腳", "營養保健", "纖體瘦身"),
    (r"小林|必理痛|鼻炎|去疣", "日本藥品", "日本 OTC／保健"),
    (r"幽門|益安寧|救心|同溢堂", "營養保健", "腸胃／心血管"),
    (r"正官庄|紅蔘|蔘", "營養保健", "滋補養生"),
    (r"足貼|去濕|臥佛", "營養保健", "祛濕排毒"),
    (r"野口|q10|輔酶", "營養保健", "輔酶 Q10"),
    (r"wyeth|materna|dha|惠氏", "營養保健", "孕婦／DHA"),
    (r"wholelove.*眼|補眼|美目", "營養保健", "護眼明目"),
    (r"寶和堂|清食|山楂", "營養保健", "綜合保健"),
    (r"雞胸肉|營美健", "食品美食", "即食食品"),
    (r"灰甲|savasia|偉哥", "個人護理", "個人護理"),
]


def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": _UA, "Accept-Language": "zh-HK,zh;q=0.9"})
    return s


def category_urls_from_sitemap(base_url: str, session: requests.Session) -> list[tuple[str, str]]:
    """回傳 [(slug, 顯示名), ...]"""
    root = base_url.rstrip("/")
    resp = session.get(f"{root}/sitemap.xml", timeout=60)
    resp.raise_for_status()
    tree = ET.fromstring(resp.content)
    out: list[tuple[str, str]] = []
    for loc in tree.findall(".//{http://www.sitemaps.org/schemas/sitemap/0.9}loc"):
        if not loc.text or "/categories/" not in loc.text:
            continue
        slug = unquote(loc.text.rstrip("/").split("/categories/")[-1])
        out.append((slug, slug))
    return out


def _extract_product_paths(html: str) -> set[str]:
    paths: set[str] = set()
    for m in re.finditer(r'["\']?/products/([^"\']+?)["\']', html):
        slug = unquote(m.group(1))
        if "{{" in slug or len(slug) < 3:
            continue
        paths.add(f"/products/{slug}")
    return paths


def scrape_category_products(
    base_url: str,
    session: requests.Session,
    categories: list[tuple[str, str]],
    workers: int = 6,
) -> dict[str, list[str]]:
    """product_path → [category_slug, ...]"""
    mapping: dict[str, set[str]] = defaultdict(set)
    root = base_url.rstrip("/")

    def fetch_cat(slug: str) -> tuple[str, set[str]]:
        paths: set[str] = set()
        for page in range(1, 51):
            url = f"{root}/categories/{slug}" + (f"?page={page}" if page > 1 else "")
            try:
                r = session.get(url, timeout=45)
                if r.status_code != 200:
                    break
                found = _extract_product_paths(r.text)
                if not found:
                    break
                before = len(paths)
                paths |= found
                if page > 1 and len(found) < 5:
                    break
                if page > 1 and len(paths) == before:
                    break
            except requests.RequestException:
                break
        return slug, paths

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futs = [pool.submit(fetch_cat, slug) for slug, _ in categories]
        for i, fut in enumerate(as_completed(futs), 1):
            slug, paths = fut.result()
            for p in paths:
                mapping[p].add(slug)
            if i % 15 == 0:
                print(f"  分類頁 {i}/{len(categories)}…", flush=True)
    return {k: sorted(v) for k, v in mapping.items()}


def _slug_label(slug: str) -> str:
    return unquote(slug.replace("-", " "))


def pick_taxonomy(slugs: list[str]) -> tuple[str, str, str]:
    """回傳 (專業分類, 詳細分類, 網站分類原文)"""
    usable = [s for s in slugs if s not in _SKIP_CATS and unquote(s) not in _SKIP_CATS]
    if not usable:
        return "", "", ""

    specific = [s for s in usable if s not in _PARENT_ONLY and unquote(s) not in _PARENT_ONLY]
    chosen = specific[0] if specific else usable[0]
    label = _slug_label(chosen)
    pro, detail = _TAXONOMY.get(chosen, _TAXONOMY.get(label, ("", label)))
    if not pro:
        pro, detail = _TAXONOMY.get(label, ("其他", label))
    site_cats = " | ".join(_slug_label(s) for s in usable[:5])
    return pro, detail, site_cats


def _match_rules(text: str, rules: list[tuple[str, str]], limit: int = 4) -> str:
    hits: list[str] = []
    for pat, label in rules:
        if re.search(pat, text, re.I):
            if label not in hits:
                hits.append(label)
        if len(hits) >= limit:
            break
    return "、".join(hits)


def classify_row(row: dict, path_to_cats: dict[str, list[str]]) -> dict[str, str]:
    url = row.get("url", "")
    path = urlparse(url).path
    name = row.get("name", "") or ""
    brand = row.get("brand", "") or ""
    desc = row.get("description", "") or ""
    blob = f"{name} {brand} {desc}".lower()

    slugs = path_to_cats.get(path, [])
    pro, detail, site_cats = pick_taxonomy(slugs)

    if not pro:
        for pat, p, d in _KEYWORD_TAXONOMY:
            if re.search(pat, f"{name} {brand} {desc}", re.I):
                pro, detail = p, d
                break
    if not pro:
        pro, detail = "未分類", "待人工覆核"

    func = _match_rules(blob, _FUNCTION_RULES)
    target = _match_rules(blob, _TARGET_RULES, limit=3)
    ptype = _match_rules(blob, _TYPE_RULES, limit=3)

    if not func and detail:
        func = detail.split("／")[0] if "／" in detail else detail

    return {
        "專業分類": pro,
        "詳細分類": detail,
        "網站分類": site_cats,
        "功能": func,
        "針對": target or "通用",
        "類型": ptype or "其他",
    }


def read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fields = list(reader.fieldnames or [])
    return fields, rows


def write_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def autosize(ws, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max_width, max((len(str(c.value or "")) for c in col), default=8) + 2)
        ws.column_dimensions[letter].width = width


def write_excel(out: Path, products: list[dict], images: list[dict], fields: list[str]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, data, cols in [
        ("產品", products, fields),
        ("圖片", images, list(images[0].keys()) if images else []),
        ("分類統計", [], []),
    ]:
        ws = wb.create_sheet(title=title)
        if title == "分類統計":
            from collections import Counter

            c_pro = Counter(r.get("專業分類", "") for r in products)
            c_detail = Counter(r.get("詳細分類", "") for r in products)
            ws.append(["專業分類", "數量"])
            for k, v in c_pro.most_common():
                ws.append([k, v])
            ws.append([])
            ws.append(["詳細分類", "數量"])
            for k, v in c_detail.most_common():
                ws.append([k, v])
            autosize(ws)
            continue
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in data:
            ws.append([row.get(c, "") for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws)
    wb.save(out)


def rebuild_zip(out: Path, xlsx: Path) -> None:
    zip_path = out / "saanfolouhk-export.zip"
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for name in ("products.csv", "images.csv", xlsx.name):
            p = out / name
            if p.is_file():
                zf.write(p, name)
        img_dir = out / "images"
        if img_dir.is_dir():
            for img in sorted(img_dir.iterdir()):
                if img.is_file():
                    zf.write(img, f"images/{img.name}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=Path("reports/saanfolouhk-export/products.csv"),
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=Path("reports/saanfolouhk-export"),
    )
    parser.add_argument("--base-url", default="https://www.saanfolouhk.com/")
    parser.add_argument("--workers", type=int, default=8)
    parser.add_argument("--skip-scrape", action="store_true", help="只用關鍵字，不抓分類頁")
    args = parser.parse_args()

    out = args.output_dir.resolve()
    products_csv = args.input.resolve()
    fields, rows = read_csv(products_csv)

    new_cols = ["專業分類", "詳細分類", "網站分類", "功能", "針對", "類型"]
    for c in new_cols:
        if c not in fields:
            fields.append(c)

    path_to_cats: dict[str, list[str]] = {}
    if not args.skip_scrape:
        session = _session()
        print("抓取網站分類對應…", flush=True)
        cats = category_urls_from_sitemap(args.base_url, session)
        path_to_cats = scrape_category_products(args.base_url, session, cats, args.workers)
        print(f"  已對應 {len(path_to_cats)} 個產品路徑", flush=True)

    for row in rows:
        row.update(classify_row(row, path_to_cats))

    write_csv(products_csv, fields, rows)
    print(f"已更新 {products_csv}", flush=True)

    images_csv = out / "images.csv"
    img_fields, img_rows = read_csv(images_csv) if images_csv.is_file() else ([], [])
    xlsx = out / "saanfolouhk-export.xlsx"
    write_excel(xlsx, rows, img_rows, fields)
    print(f"已更新 Excel → {xlsx}", flush=True)
    rebuild_zip(out, xlsx)
    print(f"已更新 ZIP → {out / 'saanfolouhk-export.zip'}", flush=True)

    from collections import Counter

    c = Counter(r["專業分類"] for r in rows)
    print("\n專業分類分佈：")
    for k, v in c.most_common():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
