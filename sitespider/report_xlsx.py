"""
Screaming Frog 風格 Excel 匯出（多工作表）。

需安裝：pip install sitespider[excel]
"""

from __future__ import annotations

from pathlib import Path

from sitespider.crawler import CrawlReport
from sitespider.issues import ISSUE_LABELS


def xlsx_available() -> bool:
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def export_xlsx(report: CrawlReport, path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError(
            "Excel 匯出需要 openpyxl，請執行：pip install 'sitespider[excel]'"
        ) from e

    wb = Workbook()
    bold = Font(bold=True)

    # --- Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    issues = report.summary_issues()
    ws_sum.append(["SiteSpider SEO 報告"])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append(["起始 URL", report.start_url])
    ws_sum.append(["模式", report.mode])
    ws_sum.append(["頁數", len(report.pages)])
    ws_sum.append(["robots 封鎖", len(report.blocked_urls)])
    ws_sum.append([])
    ws_sum.append(["問題類型", "頁數"])
    ws_sum["A7"].font = bold
    ws_sum["B7"].font = bold
    for key, urls in sorted(issues.items(), key=lambda x: -len(set(x[1]))):
        ws_sum.append([ISSUE_LABELS.get(key, key), len(set(urls))])

    # --- Internal (SF) ---
    from sitespider.report import _sf_row

    ws_int = wb.create_sheet("Internal")
    int_headers = list(_sf_row("", report.pages[next(iter(report.pages))]).keys()) if report.pages else []
    if report.pages:
        ws_int.append(int_headers)
        for c in range(1, len(int_headers) + 1):
            ws_int.cell(1, c).font = bold
        for url, p in sorted(report.pages.items()):
            row = _sf_row(url, p)
            ws_int.append([row.get(h, "") for h in int_headers])

    # --- Response Codes ---
    ws_rc = wb.create_sheet("Response Codes")
    ws_rc.append(["Address", "Status Code", "Indexability"])
    for c in range(1, 4):
        ws_rc.cell(1, c).font = bold
    for url, p in sorted(report.pages.items()):
        ws_rc.append([url, p.status, p.indexability])

    # --- Page Titles ---
    ws_pt = wb.create_sheet("Page Titles")
    ws_pt.append(["Address", "Title 1", "Title 1 Length", "Indexability"])
    for c in range(1, 5):
        ws_pt.cell(1, c).font = bold
    for url, p in sorted(report.pages.items()):
        t = p.title or ""
        ws_pt.append([url, t, len(t), p.indexability])

    # --- Pages ---
    ws_pages = wb.create_sheet("Pages")
    page_headers = [
        "url",
        "status",
        "title",
        "meta_description",
        "canonical",
        "html_lang",
        "h1",
        "word_count",
        "depth",
        "inlinks",
        "issues",
        "lh_seo",
        "response_ms",
    ]
    ws_pages.append(page_headers)
    for c in range(1, len(page_headers) + 1):
        ws_pages.cell(1, c).font = bold
    for url, p in sorted(report.pages.items()):
        lh = p.lighthouse
        ws_pages.append(
            [
                url,
                p.status,
                p.title or "",
                (p.meta_description or "")[:200],
                p.canonical or "",
                p.html_lang or "",
                " | ".join(p.h1)[:300],
                p.word_count,
                p.crawl_depth,
                len(p.inlinks),
                "; ".join(p.issues),
                lh.seo if lh else "",
                round(p.response_ms, 1),
            ]
        )

    # --- Issues ---
    ws_issues = wb.create_sheet("Issues")
    ws_issues.append(["issue", "label", "url", "status", "depth"])
    for c in range(1, 6):
        ws_issues.cell(1, c).font = bold
    for url, p in sorted(report.pages.items()):
        for issue in sorted(set(p.issues)):
            ws_issues.append(
                [issue, ISSUE_LABELS.get(issue, issue), url, p.status, p.crawl_depth]
            )

    # --- Links ---
    ws_links = wb.create_sheet("Links")
    ws_links.append(["source_url", "href", "resolved", "anchor", "type", "status", "issue"])
    for c in range(1, 8):
        ws_links.cell(1, c).font = bold
    for page_url, p in sorted(report.pages.items()):
        for link in p.links:
            ws_links.append(
                [
                    page_url,
                    link.href,
                    link.resolved,
                    link.text,
                    link.link_type,
                    link.status or "",
                    link.issue or "",
                ]
            )

    # --- Images ---
    ws_img = wb.create_sheet("Images")
    ws_img.append(["source_url", "src", "resolved", "alt", "status", "issue"])
    for c in range(1, 7):
        ws_img.cell(1, c).font = bold
    for page_url, p in sorted(report.pages.items()):
        for img in p.images:
            ws_img.append(
                [
                    page_url,
                    img.src,
                    img.resolved,
                    img.alt or "",
                    img.status or "",
                    img.issue or "",
                ]
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
